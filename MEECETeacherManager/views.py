# views.py
from django.views.generic import View, DetailView, ListView, TemplateView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import JsonResponse, FileResponse
from django.core.serializers import serialize
from django.db.models import Q, Prefetch, Min, Max
from datetime import datetime, date, time, timedelta
from django.contrib.auth.models import User 
from django.utils import timezone
import calendar
import openpyxl
from .models import (Teacher, TeacherCategory, TeacherProfile, Course, TeacherSchedule, 
                     TeacherPersonalLink, TeacherGlobalLinkURL, GlobalTeacherLink, ComplianceStatus, ComplianceTask,
                     MPA, MPASheet, MPASchedule, TimelineEvent)

from .forms import (TeacherForm, TeacherCategoryForm, TeacherProfileForm, CourseForm, TeacherScheduleForm,
                    TeacherPersonalLinkForm, TeacherGlobalLinkURLForm, GlobalTeacherLinkForm, ComplianceTaskForm,
                    MPAUploadForm, MPAEditForm, TimelineEventForm)

from .services import TimelineService 

from .utils import (
    generate_profile_pdf, 
    generate_schedule_pdf, 
    generate_compliance_report_pdf,
    generate_course_overview_pdf,
    generate_mpa_validation_pdf,   # Nueva función
    generate_intake_summary_pdf    # Nueva función
)



class TeacherDetailView(LoginRequiredMixin, DetailView):
    model = Teacher
    template_name = 'MEECETeacherManager/teacher_detail.html'
    context_object_name = 'teacher'


class TeacherManagerView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/teacher_manager.html'
    
    def get(self, request):
        form = TeacherForm()
        teachers = Teacher.objects.all().order_by('name')
        return render(request, self.template_name, {
            'form': form,
            'teachers': teachers,
        })
    
    def post(self, request):
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Docente creado exitosamente.')
            return redirect('meece_teacher:teacher_manager')
        
        teachers = Teacher.objects.all().order_by('name')
        return render(request, self.template_name, {
            'form': form,
            'teachers': teachers
        })

class TeacherUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        teacher = get_object_or_404(Teacher, pk=pk)
        form = TeacherForm(request.POST, instance=teacher)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Docente actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el docente.')
            
        return redirect('meece_teacher:teacher_manager')

class TeacherDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        teacher = get_object_or_404(Teacher, pk=pk)
        teacher.delete()
        messages.success(request, 'Docente eliminado exitosamente.')
        return redirect('meece_teacher:teacher_manager')
    


# views.py
class TeacherCategoryManagerView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/category_manager.html'
    
    def get(self, request):
        form = TeacherCategoryForm()
        categories = TeacherCategory.objects.all().order_by('name')
        return render(request, self.template_name, {
            'form': form,
            'categories': categories,
        })
    
    def post(self, request):
        form = TeacherCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('meece_teacher:category_manager')
        
        categories = TeacherCategory.objects.all().order_by('name')
        return render(request, self.template_name, {
            'form': form,
            'categories': categories
        })

class TeacherCategoryUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        category = get_object_or_404(TeacherCategory, pk=pk)
        form = TeacherCategoryForm(request.POST, instance=category)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
        else:
            messages.error(request, 'Error al actualizar la categoría.')
            
        return redirect('meece_teacher:category_manager')

class TeacherCategoryDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        category = get_object_or_404(TeacherCategory, pk=pk)
        try:
            category.delete()
            messages.success(request, 'Categoría eliminada exitosamente.')
        except ProtectedError:
            messages.error(request, 'No se puede eliminar esta categoría porque está siendo utilizada por docentes.')
        return redirect('meece_teacher:category_manager')

class TeacherProfileCreateView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/profile_form.html'

    def get(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        form = TeacherProfileForm()
        return render(request, self.template_name, {
            'form': form,
            'teacher': teacher
        })

    def post(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        form = TeacherProfileForm(request.POST)
        
        if form.is_valid():
            profile = form.save(commit=False)
            profile.teacher = teacher
            profile.save()
            messages.success(request, 'Perfil creado exitosamente.')
            return redirect('meece_teacher:teacher_manager')
            
        return render(request, self.template_name, {
            'form': form,
            'teacher': teacher
        })

class TeacherProfileUpdateView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/profile_form.html'

    def get(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        profile = get_object_or_404(TeacherProfile, teacher=teacher)
        form = TeacherProfileForm(instance=profile)
        return render(request, self.template_name, {
            'form': form,
            'teacher': teacher,
            'profile': profile
        })

    def post(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        profile = get_object_or_404(TeacherProfile, teacher=teacher)
        form = TeacherProfileForm(request.POST, instance=profile)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            return redirect('meece_teacher:teacher_manager')
            
        return render(request, self.template_name, {
            'form': form,
            'teacher': teacher,
            'profile': profile
        })
        

class CourseManagerView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/course_manager.html'
    
    def get(self, request):
        form = CourseForm()
        courses = Course.objects.all().order_by('code')
        return render(request, self.template_name, {
            'form': form,
            'courses': courses,
        })
    
    def post(self, request):
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignatura creada exitosamente.')
            return redirect('meece_teacher:course_manager')
        
        courses = Course.objects.all().order_by('code')
        return render(request, self.template_name, {
            'form': form,
            'courses': courses,
        })

class CourseUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        course = get_object_or_404(Course, pk=pk)
        form = CourseForm(request.POST, instance=course)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Asignatura actualizada exitosamente.')
        else:
            messages.error(request, 'Error al actualizar la asignatura.')
            
        return redirect('meece_teacher:course_manager')

class CourseDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        course = get_object_or_404(Course, pk=pk)
        try:
            course.delete()
            messages.success(request, 'Asignatura eliminada exitosamente.')
        except ProtectedError:
            messages.error(request, 'No se puede eliminar esta asignatura porque está siendo utilizada en cargas horarias.')
        return redirect('meece_teacher:course_manager')



class TeacherScheduleManagerView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/schedule_manager.html'
    
    def get(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        form = TeacherScheduleForm()
        schedules = TeacherSchedule.objects.filter(teacher=teacher).order_by('day', 'start_time')
        
        return render(request, self.template_name, {
            'form': form,
            'teacher': teacher,
            'schedules': schedules,
        })
    
    def post(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        form = TeacherScheduleForm(request.POST)
        
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.teacher = teacher
            schedule.save()
            messages.success(request, 'Horario agregado exitosamente.')
            return redirect('meece_teacher:teacher_schedule', teacher_id=teacher_id)
        
        schedules = TeacherSchedule.objects.filter(teacher=teacher).order_by('day', 'start_time')
        return render(request, self.template_name, {
            'form': form,
            'teacher': teacher,
            'schedules': schedules
        })

# Agregar las vistas de clonación, actualización y eliminación
class TeacherScheduleCloneView(LoginRequiredMixin, View):
    def post(self, request, pk):
        original_schedule = get_object_or_404(TeacherSchedule, pk=pk)
        
        # Crear una nueva instancia con los mismos datos
        new_schedule = TeacherSchedule.objects.create(
            teacher=original_schedule.teacher,
            course=original_schedule.course,
            day=original_schedule.day,
            start_time=original_schedule.start_time,
            end_time=original_schedule.end_time,
            start_date=original_schedule.start_date,
            end_date=original_schedule.end_date,
            modality=original_schedule.modality,
            vacancies=original_schedule.vacancies,
            intake=original_schedule.intake,
            is_accepted=False  # Por defecto no aceptado
        )
        
        messages.success(request, 'Horario clonado exitosamente. Puede editarlo según necesite.')
        return redirect('meece_teacher:teacher_schedule', teacher_id=original_schedule.teacher.id)



class TeacherScheduleUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        schedule = get_object_or_404(TeacherSchedule, pk=pk)
        form = TeacherScheduleForm(request.POST, instance=schedule)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el horario.')
        
        return redirect('meece_teacher:teacher_schedule', teacher_id=schedule.teacher.id)

class TeacherScheduleDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        schedule = get_object_or_404(TeacherSchedule, pk=pk)
        teacher_id = schedule.teacher.id
        schedule.delete()
        messages.success(request, 'Horario eliminado exitosamente.')
        return redirect('meece_teacher:teacher_schedule', teacher_id=teacher_id)
        
        
#PDF Export

class TeacherSchedulePDFView(LoginRequiredMixin, View):
    def get(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        pdf_buffer = generate_schedule_pdf(teacher)
        
        response = FileResponse(
            pdf_buffer,
            as_attachment=True,
            filename=f'carga_horaria_{teacher.name}.pdf'
        )
        return response
    
    
class TeacherProfilePDFView(LoginRequiredMixin, View):
    def get(self, request, pk):
        teacher = get_object_or_404(Teacher, pk=pk)
        pdf_buffer = generate_profile_pdf(teacher)
        
        response = FileResponse(
            pdf_buffer,
            as_attachment=True,
            filename=f'perfil_{teacher.name}.pdf'
        )
        return response
    


class GlobalTeacherLinkManagerView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/global_link_manager.html'
    
    def get(self, request):
        form = GlobalTeacherLinkForm()
        links = GlobalTeacherLink.objects.all()
        return render(request, self.template_name, {
            'form': form,
            'links': links
        })
    
    def post(self, request):
        form = GlobalTeacherLinkForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Link global creado exitosamente.')
            return redirect('meece_teacher:global_link_manager')
        
        links = GlobalTeacherLink.objects.all()
        return render(request, self.template_name, {
            'form': form,
            'links': links
        })

class GlobalTeacherLinkUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        link = get_object_or_404(GlobalTeacherLink, pk=pk)
        form = GlobalTeacherLinkForm(request.POST, instance=link)
        if form.is_valid():
            form.save()
            messages.success(request, 'Link global actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el link global.')
            
        return redirect('meece_teacher:global_link_manager')

class GlobalTeacherLinkDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        link = get_object_or_404(GlobalTeacherLink, pk=pk)
        link.delete()
        messages.success(request, 'Link global eliminado exitosamente.')
        return redirect('meece_teacher:global_link_manager')

class TeacherGlobalLinkURLManagerView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/teacher_global_links.html'
    
    def get(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        global_links = TeacherGlobalLinkURL.objects.filter(
            teacher=teacher
        ).select_related('global_link')
        
        # Crear formularios para cada link global
        forms = {
            link.id: TeacherGlobalLinkURLForm(instance=link)
            for link in global_links
        }
        
        return render(request, self.template_name, {
            'teacher': teacher,
            'global_links': global_links,
            'forms': forms
        })

class TeacherGlobalLinkURLUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        link_url = get_object_or_404(TeacherGlobalLinkURL, pk=pk)
        form = TeacherGlobalLinkURLForm(request.POST, instance=link_url)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'URL actualizada exitosamente.')
        else:
            messages.error(request, 'Error al actualizar la URL.')
            
        return redirect('meece_teacher:teacher_global_links', teacher_id=link_url.teacher.id)





class TeacherPersonalLinkCreateView(LoginRequiredMixin, View):
    def post(self, request, teacher_id):
        teacher = get_object_or_404(Teacher, id=teacher_id)
        form = TeacherPersonalLinkForm(request.POST)
        
        if form.is_valid():
            link = form.save(commit=False)
            link.teacher = teacher
            link.save()
            messages.success(request, 'Link personal creado exitosamente.')
        else:
            messages.error(request, 'Error al crear el link personal.')
            
        return redirect('meece_teacher:teacher_global_links', teacher_id=teacher_id)

class TeacherPersonalLinkUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        link = get_object_or_404(TeacherPersonalLink, pk=pk)
        form = TeacherPersonalLinkForm(request.POST, instance=link)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Link personal actualizado exitosamente.')
        else:
            messages.error(request, 'Error al actualizar el link personal.')
            
        return redirect('meece_teacher:teacher_global_links', teacher_id=link.teacher.id)

class TeacherPersonalLinkDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        link = get_object_or_404(TeacherPersonalLink, pk=pk)
        teacher_id = link.teacher.id
        link.delete()
        messages.success(request, 'Link personal eliminado exitosamente.')
        return redirect('meece_teacher:teacher_global_links', teacher_id=teacher_id)
    

# views.py
class ComplianceManagerView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/compliance_manager.html'
    
    def get(self, request):
        tasks = ComplianceTask.objects.all().order_by('-created_at')
        teachers = Teacher.objects.all().order_by('name')
        form = ComplianceTaskForm()
        
        # Preparar matriz de estados
        compliance_matrix = {}
        for teacher in teachers:
            compliance_matrix[teacher.id] = {}
            for task in tasks:
                status = ComplianceStatus.objects.get_or_create(
                    teacher=teacher,
                    task=task
                )[0]
                compliance_matrix[teacher.id][task.id] = status
        
        return render(request, self.template_name, {
            'form': form,
            'tasks': tasks,
            'teachers': teachers,
            'compliance_matrix': compliance_matrix
        })
    
    def post(self, request):
        form = ComplianceTaskForm(request.POST)
        if form.is_valid():
            task = form.save()
            messages.success(request, 'Tarea creada exitosamente.')
            return redirect('meece_teacher:compliance_manager')
        
        tasks = ComplianceTask.objects.all().order_by('-created_at')
        teachers = Teacher.objects.all().order_by('name')
        
        # Preparar matriz de estados igual que en get
        compliance_matrix = {}
        for teacher in teachers:
            compliance_matrix[teacher.id] = {}
            for task in tasks:
                status = ComplianceStatus.objects.get_or_create(
                    teacher=teacher,
                    task=task
                )[0]
                compliance_matrix[teacher.id][task.id] = status
        
        return render(request, self.template_name, {
            'form': form,
            'tasks': tasks,
            'teachers': teachers,
            'compliance_matrix': compliance_matrix
        })

class ComplianceStatusUpdateView(LoginRequiredMixin, View):
    def post(self, request):
        status_id = request.POST.get('status_id')
        action = request.POST.get('action')  # 'toggle' o 'activate'
        
        status = get_object_or_404(ComplianceStatus, id=status_id)
        
        if action == 'toggle':
            if status.is_active:
                status.is_checked = not status.is_checked
                status.save()
        elif action == 'activate':
            status.is_active = not status.is_active
            status.save()
            
        return JsonResponse({
            'success': True,
            'is_checked': status.is_checked,
            'is_active': status.is_active
        })
        
class ComplianceTaskDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = get_object_or_404(ComplianceTask, pk=pk)
        task_name = task.name
        try:
            task.delete()
            messages.success(request, f'Tarea "{task_name}" eliminada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar la tarea: {str(e)}')
        return redirect('meece_teacher:compliance_manager')
    
    
#Export pdf flujo de cumplimiento


class ComplianceReportView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/compliance_report.html'
    
    def get_task_data(self):
        """Obtiene y organiza los datos de tareas y su cumplimiento"""
        tasks = ComplianceTask.objects.all().order_by('-created_at')
        task_data = []
        
        for task in tasks:
            # Obtener todos los estados activos para esta tarea
            statuses = ComplianceStatus.objects.filter(
                task=task, 
                is_active=True
            ).select_related('teacher')
            
            completed_teachers = []
            pending_teachers = []
            
            for status in statuses:
                teacher_data = {
                    'name': status.teacher.name,
                    'rut': status.teacher.rut
                }
                
                if status.is_checked:
                    completed_teachers.append(teacher_data)
                else:
                    pending_teachers.append(teacher_data)
            
            # Ordenar las listas por nombre del profesor
            completed_teachers = sorted(completed_teachers, key=lambda x: x['name'])
            pending_teachers = sorted(pending_teachers, key=lambda x: x['name'])
            
            task_data.append({
                'name': task.name,
                'completed_teachers': completed_teachers,
                'pending_teachers': pending_teachers,
                'total_completed': len(completed_teachers),
                'total_pending': len(pending_teachers),
                'total_teachers': len(completed_teachers) + len(pending_teachers)
            })
            
        return task_data

    def get(self, request):
        """Maneja las solicitudes GET y la exportación a PDF"""
        task_data = self.get_task_data()
        
        # Si se solicita exportar a PDF
        if 'export_pdf' in request.GET:
            buffer = generate_compliance_report_pdf(task_data)
            return FileResponse(
                buffer,
                as_attachment=True,
                filename='reporte_cumplimiento.pdf'
            )
        
        # Calcular estadísticas generales
        total_tasks = len(task_data)
        total_teachers = Teacher.objects.count()
        
        context = {
            'task_data': task_data,
            'total_tasks': total_tasks,
            'total_teachers': total_teachers,
            'current_date': timezone.now()
        }
        
        return render(request, self.template_name, context)

    def post(self, request):
        """
        Maneja las solicitudes POST si necesitas agregar
        alguna funcionalidad adicional en el futuro
        """
        return redirect('meece_teacher:compliance_report')
    
    
class CourseOverviewView(LoginRequiredMixin, ListView):
    template_name = 'MEECETeacherManager/course_overview.html'
    context_object_name = 'schedules'
    
    def get_queryset(self):
        schedules = TeacherSchedule.objects.select_related(
            'course', 'teacher'
        ).order_by('teacher__name', 'start_date', 'day', 'start_time')
        
        for schedule in schedules:
            schedule.conflicts = self.get_schedule_conflicts(schedule, schedules)
            # Agregar validación contra MPAs
            schedule.mpa_conflicts = self.validate_against_mpas(schedule)
        
        return schedules

    def validate_date_weekday(self, schedule):
        """Valida que las fechas coincidan con el día de clase"""
        weekday_map = {
            'LUNES': 0, 'MARTES': 1, 'MIERCOLES': 2, 
            'JUEVES': 3, 'VIERNES': 4, 'SABADO': 5, 'DOMINGO': 6
        }
        
        # Diccionario para traducir los días
        dias_semana = {
            0: 'Lunes',
            1: 'Martes',
            2: 'Miércoles',
            3: 'Jueves',
            4: 'Viernes',
            5: 'Sábado',
            6: 'Domingo'
        }

        expected_weekday = weekday_map.get(schedule.day)
        start_weekday = schedule.start_date.weekday()
        end_weekday = schedule.end_date.weekday()
        
        if start_weekday != expected_weekday or end_weekday != expected_weekday:
            return False, (
                f"La clase está programada para {schedule.get_day_display()}, "
                f"pero inicia un {dias_semana[start_weekday]} "
                f"y termina un {dias_semana[end_weekday]}"
            )
        return True, None

    def validate_saturday_schedule(self, schedule):
        """Valida que las clases del sábado estén en el horario permitido"""
        if schedule.day == 'SABADO':
            start_limit = time(9, 0)  # 9:00
            end_limit = time(12, 0)   # 12:00
            
            if not (start_limit <= schedule.start_time <= end_limit and 
                   start_limit <= schedule.end_time <= end_limit):
                return False, (
                    "Las clases del sábado deben realizarse "
                    "entre las 09:00 y las 12:00"
                )
        return True, None

    def check_date_overlap(self, schedule1, schedule2):
        """Verifica si hay superposición entre dos rangos de fechas"""
        return not (
            schedule1.end_date < schedule2.start_date or 
            schedule2.end_date < schedule1.start_date
        )

    def get_schedule_conflicts(self, schedule, all_schedules):
        """Detecta conflictos de horario y validaciones de fechas"""
        conflicts = []
        
        # 1. Validar que el día de semana coincida con las fechas
        is_valid_date, date_message = self.validate_date_weekday(schedule)
        if not is_valid_date:
            conflicts.append({
                'type': 'error_fecha',
                'severity': 'danger',
                'message': date_message
            })

        # 2. Validar horario de sábado
        is_valid_saturday, saturday_message = self.validate_saturday_schedule(schedule)
        if not is_valid_saturday:
            conflicts.append({
                'type': 'error_horario',
                'severity': 'danger',
                'message': saturday_message
            })

        # 3. Buscar conflictos de horario
        for other in all_schedules:
            if (other.id != schedule.id and 
                other.teacher_id == schedule.teacher_id and 
                other.day == schedule.day):
                
                # Verificar superposición de fechas
                if self.check_date_overlap(schedule, other):
                    # Verificar superposición de horarios
                    times_overlap = (
                        (schedule.start_time <= other.start_time < schedule.end_time) or
                        (schedule.start_time < other.end_time <= schedule.end_time) or
                        (other.start_time <= schedule.start_time < other.end_time)
                    )
                    
                    if times_overlap:
                        conflicts.append({
                            'type': 'conflicto_horario',
                            'severity': 'warning',
                            'message': (
                                f"Existe un conflicto de horario con la asignatura {other.course.code}: "
                                f"mismo día ({schedule.get_day_display()}), "
                                f"mismo horario ({other.start_time.strftime('%H:%M')} - "
                                f"{other.end_time.strftime('%H:%M')}), y las fechas se solapan "
                                f"({other.start_date.strftime('%d/%m/%Y')} - "
                                f"{other.end_date.strftime('%d/%m/%Y')})"
                            )
                        })
        
        return conflicts

    def validate_against_mpas(self, schedule):
        """Verifica si el horario del profesor existe en alguna MPA y sus posibles conflictos"""
        # Primero filtramos por código de asignatura
        mpa_schedules = MPASchedule.objects.select_related('sheet__mpa').filter(
            course_code=schedule.course.code
        )
        
        # Normalizar el día para comparación
        schedule_day = schedule.day.upper().replace('É', 'E')  # Normalizar MIÉRCOLES a MIERCOLES
        
        # Función para normalizar horarios
        def normalize_time(time_str):
            # Remover espacios extras y convertir a minúsculas
            time_str = ' '.join(time_str.split()).lower()
            # Reemplazar cualquier separador por un espacio
            for sep in [' - ', ' a ', '-', 'a']:
                time_str = time_str.replace(sep, ' ')
            # Dividir en inicio y fin
            start, end = time_str.split()
            return f"{start} - {end}"
        
        # Convertir el horario del schedule a formato normalizado "HH:MM - HH:MM"
        schedule_time = f"{schedule.start_time.strftime('%H:%M')} - {schedule.end_time.strftime('%H:%M')}"
        schedule_time = normalize_time(schedule_time)
        
        # Agrupar schedules por INTAKE para buscar en cada uno
        intake_schedules = {}
        for mpa_schedule in mpa_schedules:
            intake = mpa_schedule.sheet.intake
            if intake not in intake_schedules:
                intake_schedules[intake] = []
            intake_schedules[intake].append(mpa_schedule)
        
        # Buscar coincidencia exacta en cada INTAKE
        for intake, schedules in intake_schedules.items():
            for mpa_schedule in schedules:
                # Normalizar el día y horario de la MPA
                mpa_day = mpa_schedule.day.upper().replace('É', 'E')
                mpa_time = normalize_time(mpa_schedule.schedule)
                
                # Verificar coincidencia exacta de todos los campos
                is_exact_match = (
                    schedule_day == mpa_day and
                    schedule_time == mpa_time and
                    schedule.start_date == mpa_schedule.start_date and
                    schedule.end_date == mpa_schedule.end_date
                )
                
                if is_exact_match:
                    # Si encontramos una coincidencia exacta, retornamos solo esa
                    return [{
                        'type': 'coincide',
                        'mpa': mpa_schedule.sheet.mpa,
                        'sheet': mpa_schedule.sheet,
                        'schedule': mpa_schedule,
                        'message': (f"Horario validado en MPA INTAKE {mpa_schedule.sheet.intake} "
                                  f"({mpa_schedule.sheet.mpa.program_name})")
                    }]
        
        # Si no encontramos coincidencia en ningún INTAKE, retornamos los INTAKEs donde se buscó
        if intake_schedules:
            return [{
                'type': 'no_encontrado',
                'message': 'No se encontró en ninguna MPA',
                'intakes': list(intake_schedules.keys())  # Lista de INTAKEs donde se buscó
            }]
        else:
            return [{
                'type': 'no_encontrado',
                'message': 'No se encontró en ninguna MPA',
                'intakes': []
            }]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_date'] = timezone.now()
        return context
    
    
    
    
class CourseOverviewPDFView(LoginRequiredMixin, View):
    def get(self, request):
        view = CourseOverviewView()
        view.request = request
        schedules = view.get_queryset()
        
        buffer = generate_course_overview_pdf(schedules)
        
        return FileResponse(
            buffer,
            as_attachment=True,
            filename='resumen_asignaturas_horarios.pdf'
        )
        


#MPA

class MPAListView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/mpa_list.html'
    
    def get(self, request):
        mpas = MPA.objects.prefetch_related('sheets', 'sheets__schedules').all()
        return render(request, self.template_name, {
            'mpas': mpas,
        })

class MPADetailView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/mpa_detail.html'
    
    def get(self, request, pk):
        mpa = get_object_or_404(MPA, pk=pk)
        return render(request, self.template_name, {
            'mpa': mpa,
        })

class MPADeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        mpa = get_object_or_404(MPA, pk=pk)
        mpa.delete()
        messages.success(request, 'MPA eliminada exitosamente.')
        return redirect('meece_teacher:mpa_list')

def parse_custom_date(date_value):
    """
    Convierte diferentes formatos de fecha al formato YYYY-MM-DD
    """
    if not date_value:
        return None
    
    if isinstance(date_value, datetime):
        return date_value.date()
    
    if isinstance(date_value, str):
        date_str = date_value.strip()
        formats = [
            '%d-%m-%Y',
            '%d.%m.%Y',
            '%d/%m/%Y',
            '%Y-%m-%d',
            '%d-%m.%Y',
        ]
        
        for date_format in formats:
            try:
                return datetime.strptime(date_str, date_format).date()
            except ValueError:
                continue
                
        try:
            parts = date_str.replace('.', '-').split('-')
            if len(parts) == 3:
                parts = [int(p) for p in parts if p.isdigit()]
                year = max(parts)
                parts.remove(year)
                day = parts[0]
                month = parts[1]
                if 1 <= day <= 31 and 1 <= month <= 12:
                    return datetime(year, month, day).date()
        except:
            pass
            
    return None

class MPAEditView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/mpa_edit.html'

    def get(self, request, pk):
        mpa = get_object_or_404(MPA, pk=pk)
        initial_data = {
            'program_name': mpa.program_name,
            'program_code': mpa.program_code,
            'faculty': mpa.faculty,
            'school': mpa.school,
            'campus': mpa.campus,
            'period': mpa.period
        }
        form = MPAEditForm(initial=initial_data)
        return render(request, self.template_name, {
            'form': form,
            'mpa': mpa
        })

    def post(self, request, pk):
        mpa = get_object_or_404(MPA, pk=pk)
        form = MPAEditForm(request.POST, instance=mpa)
        if form.is_valid():
            form.save()
            messages.success(request, 'MPA actualizada exitosamente')
            return redirect('meece_teacher:mpa_list')
        
        return render(request, self.template_name, {
            'form': form,
            'mpa': mpa
        })


class MPAUploadView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/mpa_upload.html'

    def get(self, request):
        form = MPAUploadForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = MPAUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = form.cleaned_data['file']
                workbook = openpyxl.load_workbook(excel_file)
                
                # 1. Crear el MPA base
                mpa = MPA(file_name=excel_file.name)
                header_info = {}
                
                # Buscar la información de cabecera en la primera hoja
                first_sheet = workbook[workbook.sheetnames[0]]
                for row in first_sheet.iter_rows(min_row=1, max_row=15):
                    for cell in row:
                        if cell.value:
                            value = str(cell.value).strip()
                            field_mapping = {
                                "Facultad": "faculty",
                                "Escuela": "school",
                                "Sede / Campus": "campus",
                                "Período": "period",
                                "Código de Programa": "program_code",
                                "Nombre de Programa": "program_name"
                            }
                            
                            for search_text, field_name in field_mapping.items():
                                if search_text in value:
                                    next_cell = first_sheet.cell(row=cell.row, column=cell.column + 2)
                                    header_info[field_name] = next_cell.value if next_cell.value else "No especificado"

                # Actualizar y guardar el MPA
                for field, value in header_info.items():
                    setattr(mpa, field, value)
                mpa.save()

                # 2. Procesar cada hoja
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Crear MPASheet
                    intake = sheet_name if "INTAKE" in sheet_name.upper() else f"INTAKE {sheet_name}"
                    mpa_sheet = MPASheet.objects.create(
                        mpa=mpa,
                        sheet_name=sheet_name,
                        intake=intake
                    )

                    # 3. Encontrar la tabla y sus encabezados
                    table_start_row = None
                    headers = {}
                    
                    # Buscar la fila de encabezados
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50), 1):
                        for cell in row:
                            if cell.value and "NRC" in str(cell.value):
                                table_start_row = row_idx
                                # Mapear los encabezados con sus columnas
                                for col_idx, header_cell in enumerate(row, 1):
                                    if header_cell.value:
                                        headers[col_idx] = header_cell.value.strip()
                                break
                        if table_start_row:
                            break

                    # 4. Procesar las filas de datos
                    if table_start_row:
                        for row in sheet.iter_rows(min_row=table_start_row + 1, max_row=sheet.max_row):
                            if any(cell.value for cell in row):
                                schedule_data = {
                                    'sheet': mpa_sheet,
                                    'period': '',
                                    'nrc': '',
                                    'course_code': '',
                                    'section': '',
                                    'course_name': '',
                                    'activity': '',
                                    'modality': '',
                                    'vacancies': 0,
                                    'day': '',
                                    'schedule': '',
                                    'duration': '',
                                    'start_date': None,
                                    'end_date': None,
                                    'teacher_rut': '',
                                    'teacher_name': '',
                                    'building': '',
                                    'classroom': '',
                                    'hourly_rate': None,
                                    'catalog_hours': None,
                                    'responsibility_percentage': None,
                                    'payable_hours': None,
                                    'installments': None,
                                    'installment_value': None,
                                    'total_payment': None,
                                    'folio': '',
                                    'budget': '',
                                    'minimum': ''
                                }

                                for col_idx, cell in enumerate(row, 1):
                                    header = headers.get(col_idx)
                                    if header:
                                        value = cell.value if cell.value is not None else ''
                                        
                                        field_mapping = {
                                            'Período': 'period',
                                            'NRC': 'nrc',
                                            'Código Asignatura': 'course_code',
                                            'Sección': 'section',
                                            'Nombre Asignatura': 'course_name',
                                            'Actividad': 'activity',
                                            'Modalidad (presencial, on line)': 'modality',
                                            'Cantidad de Vacantes': 'vacancies',
                                            'Día': 'day',
                                            'Horario (módulo)': 'schedule',
                                            'Duración (semanas)': 'duration',
                                            'Fecha de Inicio': 'start_date',
                                            'Fecha de Término': 'end_date',
                                            'RUT': 'teacher_rut',
                                            'Nombre Profesor': 'teacher_name',
                                            'Edificio': 'building',
                                            'Sala': 'classroom',
                                            'Valor Hora': 'hourly_rate',
                                            'Horas Catálogo': 'catalog_hours',
                                            '% Resp.': 'responsibility_percentage',
                                            '% Responsabilidad': 'responsibility_percentage',
                                            'Horas a Pagar': 'payable_hours',
                                            'Cantidad de Cuotas': 'installments',
                                            'Valor Cuota': 'installment_value',
                                            'Total a Pagar': 'total_payment',
                                            'FOLIO': 'folio',
                                            'PPTO': 'budget',
                                            'Mínimo': 'minimum'
                                        }

                                        field = field_mapping.get(header)
                                        if field:
                                            if field in ['start_date', 'end_date']:
                                                parsed_date = parse_custom_date(value)
                                                if parsed_date:
                                                    schedule_data[field] = parsed_date
                                                continue

                                            if field in ['vacancies', 'catalog_hours', 'payable_hours', 'installments']:
                                                try:
                                                    schedule_data[field] = int(float(str(value))) if value and str(value).strip() != '-' else None
                                                except (ValueError, TypeError):
                                                    schedule_data[field] = None
                                                continue

                                            if field in ['hourly_rate', 'responsibility_percentage', 'installment_value', 'total_payment']:
                                                try:
                                                    if value and str(value).strip() != '-':
                                                        clean_value = str(value).replace('$', '').replace(',', '').strip()
                                                        schedule_data[field] = float(clean_value) if clean_value else None
                                                    else:
                                                        schedule_data[field] = None
                                                except (ValueError, TypeError):
                                                    schedule_data[field] = None
                                                continue
                                            
                                            schedule_data[field] = str(value).strip() if value else ''

                                if schedule_data['course_code'] or schedule_data['teacher_name']:
                                    try:
                                        MPASchedule.objects.create(**schedule_data)
                                    except Exception as e:
                                        continue

                messages.success(request, 'MPA procesado exitosamente')
                return redirect('meece_teacher:mpa_list')

            except Exception as e:
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
                return render(request, self.template_name, {'form': form})

        return render(request, self.template_name, {'form': form})
    
    
class MPAIntakeStatusView(LoginRequiredMixin, View):
    template_name = 'MEECETeacherManager/mpa_intake_status.html'

    def get(self, request):
        today = date.today()
        intake_status = []

        # Obtener todos los MPASheet (INTAKES)
        sheets = MPASheet.objects.prefetch_related('schedules').all()

        for sheet in sheets:
            schedules = sheet.schedules.all()
            if not schedules.exists():
                continue

            # Obtener fechas globales
            start_dates = [s.start_date for s in schedules if s.start_date]
            end_dates = [s.end_date for s in schedules if s.end_date]
            
            global_start = min(start_dates) if start_dates else None
            global_end = max(end_dates) if end_dates else None

            # Buscar asignaturas actuales
            current_courses = schedules.filter(
                start_date__lte=today,
                end_date__gte=today
            ).order_by('start_date')

            # Si no hay actuales, buscar la próxima
            next_course = None
            if not current_courses.exists():
                next_course = schedules.filter(
                    start_date__gt=today
                ).order_by('start_date').first()

            # Determinar si el INTAKE ha finalizado
            is_finished = global_end and global_end < today

            intake_status.append({
                'sheet': sheet,
                'global_start': global_start,
                'global_end': global_end,
                'current_courses': current_courses,
                'next_course': next_course,
                'is_finished': is_finished,
            })

        return render(request, self.template_name, {
            'intake_status': intake_status
        })

    def post(self, request):
        sheet_id = request.POST.get('sheet_id')
        if sheet_id:
            sheet = get_object_or_404(MPASheet, id=sheet_id)
            
            # Verificar que el INTAKE haya finalizado
            last_course = sheet.schedules.order_by('-end_date').first()
            if last_course and last_course.end_date < date.today():
                sheet.delete()
                messages.success(request, f'INTAKE {sheet.intake} eliminado exitosamente')
            else:
                messages.error(request, 'No se puede eliminar un INTAKE que no ha finalizado')
        
        return redirect('meece_teacher:mpa_intake_status') 
    

class MPAValidationView(LoginRequiredMixin, View):
   template_name = 'MEECETeacherManager/mpa_validation.html'

   def get(self, request):
       validation_results = []
       mpas = MPA.objects.prefetch_related(
           'sheets', 
           'sheets__schedules'
       ).all()

       for mpa in mpas:
           mpa_validation = {
               'mpa': mpa,
               'sheets_validation': []
           }

           for sheet in mpa.sheets.all():
               sheet_validation = {
                   'sheet': sheet,
                   'conflicts': self.validate_sheet(sheet)
               }
               mpa_validation['sheets_validation'].append(sheet_validation)

           validation_results.append(mpa_validation)

       return render(request, self.template_name, {
           'validation_results': validation_results
       })

   def validate_sheet(self, sheet):
       """Valida todos los aspectos de un INTAKE/hoja"""
       conflicts = []
       schedules = sheet.schedules.all()

       if not schedules.exists():
           return []

       # 1. Validar cada horario individual
       for schedule in schedules:
           conflicts.extend(self.validate_schedule(schedule))

       # 2. Validar secuencia del INTAKE
       conflicts.extend(self.validate_intake_sequence(schedules))

       # 3. Validar conflictos de profesor
       conflicts.extend(self.validate_teacher_conflicts(schedules))

       return conflicts

   def validate_schedule(self, schedule):
       """Valida un horario individual"""
       conflicts = []

       # 1. Validar coincidencia día-fecha
       conflicts.extend(self.validate_weekday(schedule))

       # 2. Validar horarios de sábado
       conflicts.extend(self.validate_saturday_schedule(schedule))

       # 3. Validar duración en semanas
       conflicts.extend(self.validate_duration(schedule))

       return conflicts

   def validate_weekday(self, schedule):
       conflicts = []
       weekday_map = {
           'LUNES': 0, 'MARTES': 1, 'MIERCOLES': 2, 
           'JUEVES': 3, 'VIERNES': 4, 'SABADO': 5, 'DOMINGO': 6
       }
       dias_semana = {
           0: 'Lunes', 1: 'Martes', 2: 'Miércoles',
           3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'
       }

       expected_weekday = weekday_map.get(schedule.day.upper())
       if expected_weekday is not None:
           start_weekday = schedule.start_date.weekday()
           end_weekday = schedule.end_date.weekday()
           
           if start_weekday != expected_weekday:
               conflicts.append({
                   'schedule': schedule,
                   'type': 'date_mismatch',
                   'severity': 'warning',
                   'message': f'La fecha de inicio ({schedule.start_date}) es {dias_semana[start_weekday]}, pero la clase es {schedule.day}'
               })
           if end_weekday != expected_weekday:
               conflicts.append({
                   'schedule': schedule,
                   'type': 'date_mismatch',
                   'severity': 'warning',
                   'message': f'La fecha de término ({schedule.end_date}) es {dias_semana[end_weekday]}, pero la clase es {schedule.day}'
               })

       return conflicts

   def validate_saturday_schedule(self, schedule):
       conflicts = []
       if schedule.day.upper() == 'SABADO':
           try:
               start_time = datetime.strptime(schedule.schedule.split('a')[0].strip(), '%H:%M').time()
               end_time = datetime.strptime(schedule.schedule.split('a')[1].strip(), '%H:%M').time()
               
               if start_time < time(9, 0) or end_time > time(14, 0):
                   conflicts.append({
                       'schedule': schedule,
                       'type': 'invalid_saturday_schedule',
                       'severity': 'error',
                       'message': 'Las clases de sábado deben ser entre 9:00 y 14:00'
                   })
           except:
               conflicts.append({
                   'schedule': schedule,
                   'type': 'invalid_time_format',
                   'severity': 'error',
                   'message': 'Formato de horario inválido'
               })
       return conflicts

   def validate_duration(self, schedule):
       conflicts = []
       if schedule.duration:
           try:
               duration_weeks = int(schedule.duration)
               actual_weeks = (schedule.end_date - schedule.start_date).days / 7
               if abs(actual_weeks - duration_weeks) > 0.1:  # Margen de error de 1 día
                   conflicts.append({
                       'schedule': schedule,
                       'type': 'duration_mismatch',
                       'severity': 'error',
                       'message': f'La duración especificada es {duration_weeks} semanas, pero las fechas abarcan {actual_weeks:.1f} semanas'
                   })
           except ValueError:
               conflicts.append({
                   'schedule': schedule,
                   'type': 'invalid_duration',
                   'severity': 'error',
                   'message': 'Duración inválida'
               })
       return conflicts

   def validate_intake_sequence(self, schedules):
       conflicts = []
       sorted_schedules = sorted(schedules, key=lambda x: x.start_date)
       
       for i in range(len(sorted_schedules)-1):
           current = sorted_schedules[i]
           next_schedule = sorted_schedules[i+1]
           
           gap_days = (next_schedule.start_date - current.end_date).days
           if gap_days > 14:
               conflicts.append({
                   'schedule': current,
                   'related_schedule': next_schedule,
                   'type': 'large_gap',
                   'severity': 'warning',
                   'message': f'Intervalo de {gap_days} días entre {current.course_code} y {next_schedule.course_code}'
               })
       
       return conflicts

   def check_overlap(self, schedule1, schedule2):
       """Verifica si hay superposición entre dos horarios"""
       # Si no es el mismo profesor, no hay conflicto
       if schedule1.teacher_name.lower() != schedule2.teacher_name.lower():
           return False

       # Si no es el mismo día, no hay conflicto
       if schedule1.day.upper() != schedule2.day.upper():
           return False
           
       # Si no hay superposición de fechas, no hay conflicto
       if (schedule1.end_date < schedule2.start_date or 
           schedule2.end_date < schedule1.start_date):
           return False

       # Comprobar si el horario es el mismo
       try:
           schedule1_time = schedule1.schedule.replace('a', '-').strip()
           schedule2_time = schedule2.schedule.replace('a', '-').strip()
           
           # Si los horarios son diferentes, no hay conflicto
           if schedule1_time != schedule2_time:
               return False

           return True
       except:
           return False

   def validate_teacher_conflicts(self, schedules):
       conflicts = []
       
       for i, schedule1 in enumerate(schedules):
           for schedule2 in list(schedules)[i+1:]:
               if self.check_overlap(schedule1, schedule2):
                   conflicts.append({
                       'schedule': schedule1,
                       'related_schedule': schedule2,
                       'type': 'teacher_conflict',
                       'severity': 'error',
                       'message': (
                           f'Profesor {schedule1.teacher_name} tiene dos asignaturas en el mismo horario:\n'
                           f'{schedule1.course_code} ({schedule1.start_date.strftime("%d/%m/%Y")} - {schedule1.end_date.strftime("%d/%m/%Y")}) y\n'
                           f'{schedule2.course_code} ({schedule2.start_date.strftime("%d/%m/%Y")} - {schedule2.end_date.strftime("%d/%m/%Y")})\n'
                           f'Día: {schedule1.day} {schedule1.schedule}'
                       )
                   })
       return conflicts
        
        

class MPAValidationPDFView(LoginRequiredMixin, View):
    def get(self, request):
        # Obtener el filtro de la URL
        filter_type = request.GET.get('filter')
        
        validation_view = MPAValidationView()
        validation_results = []
        mpas = MPA.objects.prefetch_related(
            'sheets', 
            'sheets__schedules'
        ).all()

        for mpa in mpas:
            mpa_validation = {
                'mpa': mpa,
                'sheets_validation': [],
                'filter_type': filter_type  # Agregar el tipo de filtro
            }

            for sheet in mpa.sheets.all():
                conflicts = validation_view.validate_sheet(sheet)
                
                # Aplicar filtro si existe
                if filter_type and filter_type != 'all':
                    conflicts = [c for c in conflicts if c['type'] == filter_type]
                
                if conflicts:  # Solo incluir si hay conflictos después del filtro
                    sheet_validation = {
                        'sheet': sheet,
                        'conflicts': conflicts
                    }
                    mpa_validation['sheets_validation'].append(sheet_validation)

            if mpa_validation['sheets_validation']:  # Solo incluir si hay validaciones con conflictos
                validation_results.append(mpa_validation)
        
        pdf_buffer = generate_mpa_validation_pdf(validation_results)
        
        return FileResponse(
            pdf_buffer,
            as_attachment=True,
            filename='validacion_mpas.pdf'
        )

class MPAIntakeStatusPDFView(LoginRequiredMixin, View):
    def get(self, request):
        today = date.today()
        intake_status = []

        # Replicar la lógica de MPAIntakeStatusView
        sheets = MPASheet.objects.prefetch_related('schedules').all()
        
        for sheet in sheets:
            schedules = sheet.schedules.all()
            if not schedules.exists():
                continue

            start_dates = [s.start_date for s in schedules if s.start_date]
            end_dates = [s.end_date for s in schedules if s.end_date]
            
            global_start = min(start_dates) if start_dates else None
            global_end = max(end_dates) if end_dates else None

            current_courses = schedules.filter(
                start_date__lte=today,
                end_date__gte=today
            ).order_by('start_date')

            next_course = None
            if not current_courses.exists():
                next_course = schedules.filter(
                    start_date__gt=today
                ).order_by('start_date').first()

            is_finished = global_end and global_end < today

            intake_status.append({
                'sheet': sheet,
                'global_start': global_start,
                'global_end': global_end,
                'current_courses': current_courses,
                'next_course': next_course,
                'is_finished': is_finished,
            })
        
        pdf_buffer = generate_intake_summary_pdf(intake_status)
        
        return FileResponse(
            pdf_buffer,
            as_attachment=True,
            filename='resumen_intakes.pdf'
        )


    

class MPATimelineView(TemplateView):
    template_name = 'MEECETeacherManager/mpa_timeline.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mpa_id = self.kwargs.get('mpa_id')
        
        try:
            # Obtener el MPA y sus datos
            mpa = MPA.objects.get(id=mpa_id)
            context['mpa'] = mpa  # Importante: pasar el mpa al contexto
            timeline_data = []
            
            # Obtener todos los schedules para este MPA
            all_schedules = MPASchedule.objects.filter(
                sheet__mpa=mpa
            ).select_related('sheet').order_by('start_date')
            
            # Establecer fechas por defecto en caso de que no haya datos
            default_start = timezone.now().date()
            default_end = default_start + timedelta(days=120)  # 4 meses
            
            if not all_schedules.exists():
                context['error'] = "No hay datos de horarios para mostrar"
                context['timeline_data'] = []
                context['timeline_events'] = []
                context['years'] = [default_start.year, default_end.year]
                context['date_range'] = {
                    'start': default_start,
                    'end': default_end
                }
                context['form'] = TimelineEventForm()
                return context

            # Obtener años únicos
            years_set = set()
            for schedule in all_schedules:
                if schedule.start_date:
                    years_set.add(schedule.start_date.year)
                if schedule.end_date:
                    years_set.add(schedule.end_date.year)
            
            # Si no hay años válidos, usar el año actual
            if not years_set:
                years_set = {default_start.year}
                
            years = sorted(list(years_set))
                
            # Calcular rango de fechas
            try:
                # Intentar obtener la primera y última fecha
                first_schedule = all_schedules.first()
                last_schedule = all_schedules.last()
                
                if first_schedule and first_schedule.start_date and last_schedule and last_schedule.end_date:
                    date_range = {
                        'start': first_schedule.start_date,
                        'end': last_schedule.end_date
                    }
                else:
                    # Si alguna fecha es None, usar fechas por defecto
                    date_range = {
                        'start': default_start,
                        'end': default_end
                    }
            except (AttributeError, TypeError):
                # En caso de error, usar fechas por defecto
                date_range = {
                    'start': default_start,
                    'end': default_end
                }
            
            # Organizar datos por INTAKE
            for sheet in mpa.sheets.all():
                sheet_schedules = sheet.schedules.all().order_by('start_date')
                if sheet_schedules.exists():
                    # Filtrar horarios con fechas válidas
                    valid_schedules = []
                    for schedule in sheet_schedules:
                        if schedule.start_date and schedule.end_date:
                            valid_schedules.append({
                                'id': schedule.id,
                                'course_code': schedule.course_code,
                                'course_name': schedule.course_name,
                                'teacher_name': schedule.teacher_name,
                                'start_date': schedule.start_date,
                                'end_date': schedule.end_date,
                            })
                    
                    if valid_schedules:  # Solo agregar si hay horarios válidos
                        timeline_data.append({
                            'intake': sheet.intake,
                            'courses': valid_schedules
                        })
            
            # Obtener eventos del timeline
            timeline_events = TimelineEvent.objects.filter(
                mpa=mpa
            ).order_by('event_date')
            
            context.update({
                'timeline_data': timeline_data,
                'timeline_events': timeline_events,
                'years': years,
                'date_range': date_range,
                'form': TimelineEventForm()
            })
            
        except MPA.DoesNotExist:
            context['error'] = "MPA no encontrado"
            # Establecer valores por defecto
            context['timeline_data'] = []
            context['timeline_events'] = []
            context['years'] = [timezone.now().year]
            context['date_range'] = {
                'start': timezone.now().date(),
                'end': timezone.now().date() + timedelta(days=120)
            }
            context['form'] = TimelineEventForm()
        except Exception as e:
            context['error'] = f"Error: {str(e)}"
            # Establecer valores por defecto
            context['timeline_data'] = []
            context['timeline_events'] = []
            context['years'] = [timezone.now().year]
            context['date_range'] = {
                'start': timezone.now().date(),
                'end': timezone.now().date() + timedelta(days=120)
            }
            context['form'] = TimelineEventForm()
            
        return context

class TimelineEventCreateView(LoginRequiredMixin, View):
    def post(self, request, mpa_id):
        mpa = get_object_or_404(MPA, id=mpa_id)
        form = TimelineEventForm(request.POST)
        
        if form.is_valid():
            event = form.save(commit=False)
            event.mpa = mpa
            event.created_by = request.user
            event.save()
            
            return JsonResponse({
                'success': True,
                'event': {
                    'id': event.id,
                    'title': event.title,
                    'start': event.event_date.isoformat(),
                    'event_type': event.event_type,
                    'color': event.color,
                    'icon': event.icon
                }
            })
        
        return JsonResponse({
            'success': False,
            'errors': form.errors
        })

class TimelineEventUpdateView(LoginRequiredMixin, View):
    def post(self, request, mpa_id, event_id):
        mpa = get_object_or_404(MPA, id=mpa_id)
        event = get_object_or_404(TimelineEvent, id=event_id, mpa=mpa)
        
        # Obtener datos del formulario
        title = request.POST.get('title')
        event_type = request.POST.get('event_type')
        event_date = request.POST.get('event_date')
        color = request.POST.get('color')
        icon = request.POST.get('icon')
        
        if not all([title, event_type, event_date]):
            return JsonResponse({
                'success': False,
                'errors': {
                    'title': ['Este campo es obligatorio'] if not title else [],
                    'event_type': ['Este campo es obligatorio'] if not event_type else [],
                    'event_date': ['Este campo es obligatorio'] if not event_date else []
                }
            })
        
        # Actualizar el evento
        event.title = title
        event.event_type = event_type
        event.event_date = event_date
        event.color = color
        event.icon = icon
        event.save()
        
        return JsonResponse({
            'success': True,
            'event': {
                'id': event.id,
                'title': event.title,
                'start': event.event_date.isoformat(),
                'event_type': event.event_type,
                'color': event.color,
                'icon': event.icon
            }
        })

class TimelineEventDeleteView(LoginRequiredMixin, View):
    def post(self, request, mpa_id, event_id):
        mpa = get_object_or_404(MPA, id=mpa_id)
        event = get_object_or_404(TimelineEvent, id=event_id, mpa=mpa)
        
        # Eliminar el evento
        event.delete()
        
        return JsonResponse({
            'success': True
        })